"""Scratch extraction script (not part of the app) — reads SRM_Full_Workbook_Real_Data.xlsx's
'SRM' tab and writes src/data/sandRiverFinance.json. Run once, then delete."""
import json
import openpyxl

wb = openpyxl.load_workbook('SRM_Full_Workbook_Real_Data.xlsx', data_only=True)
ws = wb['SRM']

PERIODS = {'m': 'D', 'y': 'P', 'a': 'Y'}
BUDGET_COLS = {'m': 'F', 'y': 'Q', 'a': 'AA'}
LY_COLS = {'m': 'G', 'y': 'R', 'a': 'AB'}


def cell(row, col_letter):
    from openpyxl.utils import column_index_from_string
    v = ws.cell(row=row, column=column_index_from_string(col_letter)).value
    return float(v) if v is not None else None


def rdol(v):
    return None if v is None else round(v)


def rag(actual, budget, last_year):
    if actual is None or budget is None:
        return None
    if actual >= budget:
        return 'green'
    if budget != 0 and actual >= budget * 0.95:
        return 'amber'
    if last_year is not None and actual >= last_year:
        return 'red'
    return 'deepRed'


KPI_ROWS = {'netRevenue': 111, 'contributionToHO': 15, 'ebitda': 18, 'netProfit': 23}

kpis = {}
for period, acol in PERIODS.items():
    bcol = BUDGET_COLS[period]
    lcol = LY_COLS[period]
    period_kpis = {}
    for key, row in KPI_ROWS.items():
        actual = cell(row, acol)
        budget = cell(row, bcol)
        last_year = cell(row, lcol)
        variance_abs = rdol(actual - budget) if (actual is not None and budget is not None) else None
        variance_pct = round((actual - budget) / abs(budget) * 1000) / 10 if budget not in (None, 0) else None
        period_kpis[key] = {
            'status': 'ok',
            'value': rdol(actual),
            'budget': rdol(budget),
            'lastYear': rdol(last_year),
            'varianceAbs': variance_abs,
            'variancePct': variance_pct,
            'rag': rag(actual, budget, last_year),
        }
    kpis[period] = period_kpis

# plLines — (key, row, sign) where sign=1 keeps the sheet's own sign (cost rows already negative
# in the top Income Statement summary), sign=1 for positive-magnitude detail/subtotal rows too
# (sheet already stores them positive) — no flipping needed anywhere, kept explicit for clarity.
PL_ROWS = [
    ('grossRevenue', 'Total Gross Revenue', 'revenue', False, 87),
    ('netRoomRevenue', 'Total Net Room Revenue', 'revenue', True, 98),
    ('netExtraSales', 'Total Net Extra Sales', 'revenue', True, 99),
    ('netRevenue', 'Total Net Revenue', 'revenue', False, 111),
    ('variableCosts', 'Variable Costs Sub Total', 'costs', True, 135),
    ('staffCosts', 'Total Staff Cost', 'costs', True, 165),
    ('fixedCosts', 'Fixed Cost Sub Total', 'costs', True, 186),
    ('otherCosts', 'Other Cost Sub Total', 'costs', True, 223),
    ('managedCosts', 'Managed Costs', 'costs', False, 13),
    ('imposedCosts', 'Imposed Costs', 'costs', False, 14),
    ('contribution', 'Contribution', 'summary', False, 15),
    ('hoCosts', 'HO Costs', 'summary', False, 17),
    ('ebitda', 'EBITDA', 'summary', False, 18),
    ('da', 'D&A', 'summary', False, 19),
    ('financeCost', 'Finance Cost', 'summary', False, 21),
    ('nop', 'NOP', 'summary', False, 23),
]

pl_lines = []
for key, label, section, detail_only, row in PL_ROWS:
    value = {}
    budget = {}
    variance = {}
    for period, acol in PERIODS.items():
        bcol = BUDGET_COLS[period]
        a = cell(row, acol)
        b = cell(row, bcol)
        value[period] = rdol(a)
        budget[period] = rdol(b)
        variance[period] = rdol(a - b) if (a is not None and b is not None) else None
    pl_lines.append({
        'key': key, 'label': label, 'section': section, 'detailOnly': detail_only,
        'status': 'ok', 'value': value, 'budget': budget, 'variance': variance,
    })

# Top driver placeholders — never sourced from this sheet, stay NDL/TBD.
for i in (1, 2, 3):
    pl_lines.append({
        'key': f'topDriver{i}', 'label': f'Top revenue/cost driver #{i} (TBD)', 'section': 'drivers',
        'detailOnly': True, 'status': 'ndl', 'value': None, 'budget': None, 'variance': None,
    })

# Monthly series — calendar Jan-Dec 2026. Budget: BK-BV. Actual: CM-CX (real Jan-Jun 2026,
# null for Jul-Dec — those months haven't happened yet, not $0).
from openpyxl.utils import column_index_from_string


def monthly(row, start_col, n=12):
    c0 = column_index_from_string(start_col)
    return [ws.cell(row=row, column=c0 + i).value for i in range(n)]


MONTH_LABELS = [f'{m} 2026' for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']]
CURRENT_MONTH = 6  # D3 = 2026-06-01

monthly_budget = {
    'months': MONTH_LABELS,
    'revenue': [rdol(v) for v in monthly(111, 'BK')],
    'managedCosts': [rdol(v) for v in monthly(13, 'BK')],
    'imposedCosts': [rdol(v) for v in monthly(14, 'BK')],
}


def actual_or_null(vals):
    return [rdol(v) if (v is not None and i < CURRENT_MONTH) else None for i, v in enumerate(vals)]


monthly_actual = {
    'months': MONTH_LABELS,
    'revenue': actual_or_null(monthly(111, 'CM')),
    'managedCosts': actual_or_null(monthly(13, 'CM')),
    'imposedCosts': actual_or_null(monthly(14, 'CM')),
}

# Narrative — deterministic template from real YTD (+ MTD for the occupancy/cost-per-$1 pills)
# numbers, same pattern as execNarrative.ts's buildExecutiveNarrative.
y = kpis['y']
m_occ = cell(7, 'D')
m_cost = cell(260, 'D')
m_rev = cell(111, 'D')
y_ebitda_gap = rdol(cell(18, 'P') - cell(18, 'Q'))
y_rev_vs_ly_pct = round((cell(111, 'P') - cell(111, 'R')) / cell(111, 'R') * 1000) / 10

def fmtk(v):
    """Magnitude-only $K formatter for use alongside a direction word (loss/profit, miss/beat) —
    avoids a double-negative like 'a -$263.7K loss'."""
    return f"${abs(v)/1000:.1f}K"


narrative = {
    'headline': (
        f"June YTD Net Revenue is running {abs(y['netRevenue']['variancePct'])}% "
        f"{'behind' if y['netRevenue']['variancePct'] < 0 else 'ahead of'} budget, "
        f"but Net Profit has swung to a {fmtk(y['netProfit']['value'])} "
        f"{'loss' if y['netProfit']['value'] < 0 else 'profit'} — "
        f"a {fmtk(y['netProfit']['varianceAbs'])} miss vs a "
        f"${y['netProfit']['budget']/1000:.1f}K budgeted profit."
    ),
    'body': (
        f"Year-to-date Net Revenue of ${y['netRevenue']['value']/1e6:.2f}M is "
        f"{abs(y_rev_vs_ly_pct)}% {'ahead of' if y_rev_vs_ly_pct >= 0 else 'behind'} the same period "
        f"last year, even while running {abs(y['netRevenue']['variancePct'])}% short of budget. Further "
        f"down the P&L, a Head Office cost charge with zero budget this year "
        f"(${abs(pl_lines[11]['value']['y'])/1000:.1f}K actual vs $0 budgeted) is the single largest "
        f"driver pulling EBITDA to a {fmtk(y['ebitda']['value'])} "
        f"{'loss' if y['ebitda']['value'] < 0 else 'profit'} (a {fmtk(y['ebitda']['varianceAbs'])} "
        f"miss vs budget) and Net Profit to a loss — both now behind budget AND last year, per the RAG "
        f"classification. Re-extracted directly from the SRM tab of the June 2026 MIS workbook; not yet "
        f"independently reviewed."
    ),
    'pills': [
        {'label': 'June Occupancy', 'value': f'{m_occ*100:.1f}%'},
        {'label': 'Cost per $1 Revenue (MTD)', 'value': f'${m_cost/m_rev:.2f}'},
        {'label': 'YTD EBITDA Gap to Budget', 'value': f'-${abs(y_ebitda_gap)/1000:.0f}K' if y_ebitda_gap < 0 else f'+${y_ebitda_gap/1000:.0f}K'},
    ],
}

data = {
    'property': {
        'name': 'Sand River Mara Eco Camp',
        'propertyId': 'WB640',
        'operatorLabel': 'Operated by Elewana Collection',
    },
    'reportPeriod': {'label': 'Jun', 'month': 6, 'year': 2026},
    'kpis': kpis,
    'narrative': narrative,
    'charts': {
        'cumulativeRevenuePace': 'ok',
        'monthlyCostStack': 'ok',
        'netProfitWaterfall': 'ok',
    },
    'monthlyBudget': monthly_budget,
    'monthlyActual': monthly_actual,
    'plLines': pl_lines,
}

with open('sandRiverFinance.json', 'w') as f:
    json.dump(data, f, indent=2)

print('Wrote sandRiverFinance.json')
print('Headline:', narrative['headline'])
print('Body:', narrative['body'])
print('Pills:', narrative['pills'])
